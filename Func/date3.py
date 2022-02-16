from openpyxl import Workbook, load_workbook

class Date3(object):

    def __init__(self):

        wb = load_workbook('Xlsx/task_support.xlsx', data_only=True)
        self.ws = wb.active
        self.cell_g = []
        self.lust_week_day = 0

        for cell in self.ws['G']:
            self.cell_g.append(cell.value)

        del self.cell_g[0:2]

        """
            Здесь я приблизительно вычислил последние вторники месяца. Я не брал все даты, брал только те,
            что являлись уже вторником, и пускал их в проверку, а так же на некоторых проверках 
            брал среднее число среди группы месяцев.        
        
        """

        for i in self.cell_g:         
            if len(i) != 0:
                y = int(i[-2:])
                m = int(i[:2])
                d = int(i[3:5])
                if m == 1 or 10:
                    cod = (6 + y + y/4) % 7     
                    date = (d + 1 + cod) % 7
                    self.diff = 31 - d          #подсчёт дне до конца месяца
                    if date == 3:
                        if self.diff < 7:            #Проверка на конец месяца
                            self.lust_week_day += 1
                elif m == 5:
                    cod = (6 + y + y/4) % 7
                    date = (d + 2 + cod) % 7
                    self.diff = 31 - d
                    if date == 3:
                        if self.diff < 7:
                            self.lust_week_day += 1
                elif m == 8:
                    cod = (6 + y + y/4) % 7
                    date = (d + 3 + cod) % 7
                    self.diff = 31 - d
                    if date == 3:
                        if self.diff < 7:
                            self.lust_week_day += 1
                elif m == 2 or 3 or 11:
                    cod = (6 + y + y/4) % 7
                    date = (d + 4 + cod) % 7
                    self.diff = 29 - d
                    if date == 3:
                        if self.diff < 7:
                            self.lust_week_day += 1
                elif m == 6:
                    cod = (6 + y + y/4) % 7
                    date = (d + 5 + cod) % 7
                    self.diff = 30 - d
                    if date == 3:
                        if self.diff < 7:
                            self.lust_week_day += 1
                elif m == 9 or 12:
                    cod = (6 + y + y/4) % 7
                    date = (d + 6 + cod) % 7
                    self.diff = 30 - d
                    if date == 3:
                        if self.diff < 7:
                            self.lust_week_day += 1
                elif m == 4 or 7:
                    cod = (6 + y + y/4) % 7
                    date = (d + cod) % 7
                    self.diff = 30 - d
                    if date == 3:
                        if self.diff < 7:
                            self.lust_week_day += 1

if __name__ == "__main__":
    print(Date3().lust_week_day)