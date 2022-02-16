from openpyxl import Workbook, load_workbook

class Date2(object):

    def __init__(self):

        wb = load_workbook('Xlsx/task_support.xlsx', data_only=True)
        self.ws = wb.active
        self.cell_f = []
        self.week_day = 0

        for cell in self.ws['F']:
            self.cell_f.append(cell.value)

        del self.cell_f[0:2]

        for i in self.cell_f:
            if len(i) != 0:
                y = i[:4]            #Отделяю М%:Д%:Г%:
                m = int(i[5:7])
                d = int(i[8:10])
                y = int(y[-2:])
                if m == 1 or 10:                #Проверка на месяц
                    cod = (6 + y + y/4) % 7     #Код года
                    date = (d + 1 + cod) % 7    #формула дня недели
                    if date == 3:               #Проверка дня недели на "Вторник"
                        self.week_day += 1
                elif m == 5:
                    cod = (6 + y + y/4) % 7
                    date = (d + 2 + cod) % 7
                    if date == 3:
                        self.week_day += 1
                elif m == 8:
                    cod = (6 + y + y/4) % 7
                    date = (d + 3 + cod) % 7
                    if date == 3:
                        self.week_day += 1
                elif m == 2 or 3 or 11:
                    cod = (6 + y + y/4) % 7
                    date = (d + 4 + cod) % 7
                    if date == 3:
                        self.week_day += 1
                elif m == 6:
                    cod = (6 + y + y/4) % 7
                    date = (d + 5 + cod) % 7
                    if date == 3:
                        self.week_day += 1
                elif m == 9 or 12:
                    cod = (6 + y + y/4) % 7
                    date = (d + 6 + cod) % 7
                    if date == 3:
                        self.week_day += 1
                elif m == 4 or 7:
                    cod = (6 + y + y/4) % 7
                    date = (d + cod) % 7
                    if date == 3:
                        self.week_day += 1

if __name__ == "__main__":
    print(Date2().week_day)


    