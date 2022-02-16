from openpyxl import Workbook, load_workbook

class Date1(object):

    def __init__(self):
     
        wb = load_workbook('Xlsx/task_support.xlsx', data_only=True)
        self.ws = wb.active
        self.cell_e = []
        self.day_week = 0

        for cell in self.ws['E']:
            self.cell_e.append(cell.value)

        del self.cell_e[0:2]

        for i in self.cell_e:            #Ищу день недели по признаку, в начале строки
            if i.startswith('Tue'):
                self.day_week += 1

               

if __name__ == "__main__":
    
    print(Date1().day_week)