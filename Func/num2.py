from openpyxl import Workbook, load_workbook

class Num2(object):

    def __init__(self):

        wb = load_workbook('Xlsx/task_support.xlsx', data_only=True)
        self.ws = wb.active
        self.cell_c = []
        self.long = []
        
        for cell in self.ws['C']:
            self.cell_c.append(cell.value)

        del self.cell_c[0:2]

        for i in self.cell_c:                           #Выявляю простые числа
            for i in range(2, len(self.cell_c)):
                for j in self.long:
                    if i % j == 0:
                        break
                else:
                    self.long.append(i)
        self.length = len(self.long)            
    


if __name__ == "__main__":
    print(Num2().length)       
