from openpyxl import Workbook, load_workbook

class Num1(object):

    def __init__(self):

        wb = load_workbook('Xlsx/task_support.xlsx', data_only=True)
        self.ws = wb.active
        self.cell_b = []
        self.even = 0

        for cell in self.ws['B']:           #Проверяю колонку В, и переношу в список
            self.cell_b.append(cell.value)

        del self.cell_b[0:2]          #убираю лишнее из списка

        for i in self.cell_b:         #Вычисление чётных чисел
            if i % 2 == 0:
                self.even += 1

if __name__ == "__main__":
    print(Num1().even)
        