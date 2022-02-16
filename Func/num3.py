from openpyxl import Workbook, load_workbook
import decimal

class Num3(object):

    def __init__(self):
        decimal.getcontext().prec = 1

        self.wb = load_workbook('Xlsx/task_support.xlsx', data_only=True)
        self.ws = self.wb.active
        self.cell_d = []
        self.compare = 0

        for cell in self.ws['D']:
            self.cell_d.append(cell.value)

        del self.cell_d[0:2]

if __name__ == "__main__":      #Не смог придумать решение
    print(Num3().compare)        

        

        
