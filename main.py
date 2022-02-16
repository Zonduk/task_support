from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import Func.num1
import Func.num2
import Func.num3
import Func.date1
import Func.date2
import Func.date3

class Main(object):

    def __init__(self):
        
        self.wb = load_workbook('Xlsx/result.xlsx')
        self.ws = self.wb.active
        self.ws.title = 'Base'

        self.head = ['Имена'] + ['num1'] + ['num2'] + ['num3'] + ['date1'] + ['date2'] + ['date3']
        self.ws.append(self.head)         

        self.count_answer = [' ',Func.num1.Num1().even, Func.num2.Num2().length, ' ', Func.date1.Date1().day_week, Func.date2.Date2().week_day, '~'+str(Func.date3.Date3().lust_week_day)]
        
        for col in range(1):
            self.ws.append(self.count_answer)
        
        for col in range(1, 8):
            self.ws[get_column_letter(col) + '1'].font = Font(bold=True)    

        


if __name__ == "__main__":
    Main().wb.save('Xlsx/result.xlsx')
    